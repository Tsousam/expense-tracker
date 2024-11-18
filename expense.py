"""
Expense Tracker
Tiago Martins
Github - tsousam
edX - tmartinss
Porto, Portugal
"""

import calendar
import datetime
import glob
import openpyxl
import os
import random
import shutil
import sys
from tabulate import tabulate
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill



def main():
    clear_terminal()

    # Get current year to fetch .xls from that year.
    current_date = datetime.datetime.now()
    date = current_date.date()
    year = date.strftime("%Y")

    file_name = None

    while True:
        try:

            clear_terminal()

            print("\nExpense Tracker")

            # Menu
            print("""\n1. Create New File \n2. Insert Data \n3. View Data \n4. Erase Data  \n5. Merge Yearly Expenses \n6. Export File \n0. Exit\n""")

            menu_option = validate_menu_option()

            match menu_option:
                # 1. Create File
                case 1:
                    file_name = create_file()

                    other_operation_prompt()
                    continue


                # 2. Insert Data
                case 2:
                    file_name = open_or_export_file("open", True)

                    month_number, month_name = select_month("insert")

                    add_new_sheet(file_name, month_name, False)

                    rearrange_sheets(file_name)

                    insert_expenses(file_name, month_name, month_number, year)

                    other_operation_prompt()
                    continue


                # 3. View Data
                case 3:
                    file_name = open_or_export_file("open", True)

                    month_name = month_picker(file_name, "view")

                    view_data(file_name, month_name, False)

                    other_operation_prompt()
                    continue


                # 4. Erase Data
                case 4:
                    file_name = open_or_export_file("open", False)

                    month_name = month_picker(file_name, "erase")

                    erase_data(file_name, month_name)

                    other_operation_prompt()
                    continue


                # 5. Merge Yearly Expenses
                case 5:

                    total_sheet = "Total"

                    file_name = open_or_export_file("open", False)

                    erase_total_sheet(file_name, total_sheet)

                    add_new_sheet(file_name, total_sheet, False)

                    rearrange_sheets(file_name)

                    dict_merger = get_yearly(file_name)

                    merge_yearly(file_name, total_sheet, dict_merger)

                    view_data(file_name, "Total", True)

                    other_operation_prompt()
                    continue

                # 6. Export File
                case 6:
                    file_name = open_or_export_file("export", False)

                    export_file(file_name)

                    other_operation_prompt()
                    continue

                # 0. Exit
                case 0:
                    break


                # Other. Exit
                case _:
                    break

            #break

        except KeyboardInterrupt:
            clear_terminal()

        sys.exit("Exiting Expense Tracker...\n")
        


def add_new_sheet(file_name, sheet_name, messages):   

    wb = load_workbook(file_name)

    # If desired sheet doesn't exist
    if not sheet_name in wb.sheetnames:
        ws = wb.create_sheet(title=sheet_name)

        # New sheet formatting
        if sheet_name != "Total":
            # Inserting data
            ws.append(['Date', 'Description', 'Amount', 'Category', 'Total'])
            font = Font(#color='00FF0000', 
                bold=True)
            alignment = Alignment(horizontal='center', vertical='center')

            color_list = ["FF4974A5", "FF03989E", "FF843C54", "FF6FA287", "FFD59890"]

            fill = PatternFill(start_color=random.choice(color_list), end_color=random.choice(color_list), fill_type='solid')
        else:
            ws.append(['Category', 'Quantity', 'Amount', 'Total'])
            font = Font(#color='00FF0000', 
                bold=True)
            alignment = Alignment(horizontal='center', vertical='center')

            fill = PatternFill(start_color="FFE79C2A", end_color="FFE79C2A", fill_type='solid')

        for cell in ws[1:1]:
            cell.font = font
            cell.alignment = alignment
            cell.fill = fill

        ws.auto_filter.ref = ws.dimensions
        
        if messages == True:
            print(f"\nThere was no sheet of {sheet_name}. A new one was created.")

    # Delete the default sheet named "Sheet".
    if "Sheet" in wb.sheetnames:
        del wb['Sheet']
    wb.save(file_name)



# Simply clear any OS Terminal
def clear_terminal():

    os.system('cls' if os.name=='nt' else 'clear')



# Menu Create File Option
def create_file():

    clear_terminal()
    while True:
        try:
            # 2024
            year = str(input("\nFrom what year is the file?\n")).strip()
            file_name = f"Expenses{year}.xlsx"

            if os.path.exists(file_name):
                replace_workbook = str(input(f"\nThere is a file already created for {year}. Do you want to replace it? y/n\n")).lower()
                if replace_workbook == "n":
                    clear_terminal()
                    return None
                elif replace_workbook != "n" and replace_workbook != "y":
                    print("Invalid input. Please enter A or n.")
                    continue
            
            # Create a new workbook
            wb = Workbook()

            # Save the workbook
            wb.save(file_name)

            # Remove extension to present file name to the user
            temp_name = file_name.replace('.xlsx', '')

            print(f"\nCreated a new file: {temp_name}")

            return file_name
        except ValueError:
            print("\n### Please insert a valid year. ###\n")
            continue

        except KeyboardInterrupt:
            clear_terminal()
            return None
        return None
    


def erase_data(file_name, month_name):

    wb = load_workbook(file_name)

    ws = wb[month_name]

    clear_terminal()
    print("\n### ERASER ###")

    data =[]

    # Use enumerate to handle the row index automatically
    # Starting from second row
    for index, row in enumerate(ws.iter_rows(min_row=2), start=1): 
        row_values = [f"Row {index}"]

        for cell in row:
            if isinstance(cell, openpyxl.cell.cell.MergedCell):
                continue
            row_values.append(cell.value)  

         # Check if the row list is not empty
        if row_values: 
            # Append the row list to the main data list
            data.append(row_values)  

    header = ["Day", "Description", "Amount", "Category", "Total"]
    print("")
    print(tabulate(data, header, tablefmt="grid"))

    row = int(input("\nWhich row do you want to erase?\n"))

    ws.delete_rows(row+1)

    total = 0

    for column in ws.iter_cols(): 
        # Get the value of the first cell in the column (the cell with the column name) 
        column_name = column[0].value 
        if column_name == "Amount": 
            for cell in column[1:]: 
                #print(cell.value)
                total += float(cell.value)

    ws['E2'] = total
    # Apply number format tfor two decimal places
    ws["E2"].number_format = '0.00'

    clear_terminal()
    print(f"\nRow {row} has been erased.")

    wb.save(file_name)



def erase_total_sheet(file_name, total_sheet):

    wb = load_workbook(file_name)

    if total_sheet in wb.sheetnames:
        del wb[f'{total_sheet}']
    wb.save(file_name)



def export_file(file_name):

    if file_name == None:
        sys.exit("There was a problem obtaining the file's name.")
       
    # Save the workbook and copy to user's Desktop
    desktop = os.path.normpath(os.path.expanduser("~/Desktop"))
    shutil.copy2(file_name, f'{desktop}/{file_name}')
    print(f"\n{file_name} was exported to your Desktop.")



def get_yearly(file_name):

    wb = load_workbook(file_name)

    dict_merger = {}

    for month_name in wb.sheetnames:
        if month_name == "Total":
            continue

        ws = wb[month_name]

        # If sheet is empty, skip it
        if ws['C2'].value is None:
            continue
    
        for index in range(2, ws.max_row+1):
            key_category = ws[f'D{index}'].value
            value_ammount = ws[f'C{index}'].value

            if key_category not in dict_merger:
                dict_merger[key_category] = [1, value_ammount]
            else:
                dict_merger[key_category][0] += 1
                dict_merger[key_category][1] += value_ammount

    return dict_merger



def insert_expenses(file_name, month_name, month_number, year):

    wb = load_workbook(file_name)

    ws = wb[month_name]

    # Insert expenses until the user CTRL+C the Program.
    counter = 1
    while True:
        try:
            clear_terminal()
            print(f"\nInsert Expense #{counter}")
            day = int(input("\nDay(1-31): \n"))

            if not day in range (1, 31):
                print("\nDay must be between 1 and 31.")
                continue

            description = str(input("\nDescription: \n")).strip()
            amount = float(input("\nAmount: \n"))
            category = str(input("\nCategory: \n")).strip()

            date = f"{day}/{month_number}/{year}"
            #amount = float('{:.2f}'.format(amount))
            list = [date, description, amount, category]

            ws.append(list)

            ws.column_dimensions["A"].width = 17
            ws.column_dimensions["B"].width = 45
            ws.column_dimensions["C"].width = 15
            ws.column_dimensions["D"].width = 30
            ws.column_dimensions["E"].width = 15

            for cell in ws['C']:
                cell.number_format = '0.00'

            wb.save(file_name)
            counter += 1
            clear_terminal()
            print("\n### INSERTER ###")
        except ValueError:
            print("\n### Please insert correct data ###\n")
            continue

        except KeyboardInterrupt:
            total = 0

            for column in ws.iter_cols(): 
                column_name = column[0].value 
                # Check if the column is the "Name" column 
                if column_name == "Amount": 
                    for cell in column[1:]: 
                        #print(cell.value)
                        total += float(cell.value)

            ws['E2'] = total
            ws["E2"].number_format = '0.00'

            wb.save(file_name)
        
            clear_terminal()

            print("\nEntries were saved.")
            break



def merge_yearly(file_name, total_sheet, dictionary):

    if dictionary == False:
        raise ValueError("Cannot merge expenses. Dictionary is empty.")
    
    wb = load_workbook(file_name)
    
    ws = wb[total_sheet]

    for key, value in dictionary.items():
        row_data = [key, value[0], value[1]]
        ws.append(row_data)

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15
    for cell in ws['C']:
        cell.number_format = '0.00'


    wb.save(file_name)

    clear_terminal()
    
    total = 0

    for column in ws.iter_cols(): 
        column_name = column[0].value 
        if column_name == "Amount": 
            for cell in column[1:]: 
                #print(cell.value)
                total += float(cell.value)

    ws['D2'] = total
    # Apply number format to ensure two decimal places
    ws["D2"].number_format = '0.00'

    wb.save(file_name)



def month_picker(file_name, action):

    wb = load_workbook(file_name)

    temp_name = file_name.replace('.xlsx', '')
    clear_terminal()
    print(f"\n{temp_name} loaded successfully.")

    print(f"\n### {action.upper()}ER ###")

    print(f"\nWhich month would you like to {action} data from?\n")

    list_months_available = wb.sheetnames

    for index in range(len(list_months_available)):
        print(f"{index+1}. {list_months_available[index]}")

    while True:
        try:
            print("\n")
            month_index = int(input(""))

            print(len(list_months_available))
            if not month_index in range (1, len(list_months_available)+1):
                print(f"\nPlease select one of the options available. (1-{len(list_months_available)})")
                continue
            break

        except ValueError:
            print(f"\nYou need to choose a number from the options available. (1-{len(list_months_available)})")
            pass    

        except KeyboardInterrupt:
            os.system('cls' if os.name=='nt' else 'clear')
            break

    if not 'month_index' in locals():
        raise ValueError("The month chosen does not exist.")            
    
    month_name = list_months_available[month_index-1]

    return month_name



def open_or_export_file(action, messages):

    clear_terminal()
    
    print(f"\nWhich file would you like to {action}?\n")

    # Fetch all .xlsx files
    file_list = glob.glob("*.xlsx")

    # Print menu of file list with removed files extension
    for index in range(len(file_list)):
        print(f"{index+1}. {file_list[index].replace('.xlsx', '')}")

    while True:
        try:
            print("")
            file_index = int(input(""))

            if not file_index in range (1, len(file_list)+1):
                print(f"\nPlease select one of the options available. (1-{len(file_list)})")
                continue
            break

        except ValueError:
            print(f"\nYou need to choose a number from the options available. (1-{len(file_list)})")
            pass     

        except KeyboardInterrupt:
            clear_terminal()
            break    

    if not 'file_index' in locals():
        raise ValueError("The file chosen does not exist.")           

    file_name = file_list[file_index-1]

    # Load the chosen file
    wb = load_workbook(file_name)
    temp_name = file_name.replace('.xlsx', '')
    clear_terminal()
    if action == "open":
        if messages == True:
            print(f"\n{temp_name} loaded successfully.")

    return file_name

    #pattern = r"^Expenses\d{1-4}"
    #match = re.search(pattern, temp_name)

    #if match:
        #year = match.group(1)

        #return wb, file_name, year
    #else:
        #sys.exit("Failure finding the year.")



def other_operation_prompt():

    while True:
        try:
            prompt = str(input("\nDo you want to do any other action? y/n \n")).strip().lower()

            if prompt == "y":
                clear_terminal()
                return
            elif prompt == "n":
                clear_terminal()
                sys.exit("Exiting Expense Tracker...\n")
            else:
                print("\nInvalid input. Please enter y or n.")
                continue
        except KeyboardInterrupt:
            clear_terminal()
            return

   

def rearrange_sheets(file_name):

    wb = load_workbook(file_name)
    # Get the current sheet names.
    sheet_names = wb.sheetnames

    months = [datetime.datetime.strptime(str(i), "%m").strftime("%B") for i in range(1, 13)]

    # We need to separate the monthly sheets from non-monthly ones
    monthly_sheets = []
    for name in sheet_names:
        if name in months:
            monthly_sheets.append(name)

    if "Total" in sheet_names:
        total_sheet = "Total"
    else:
        total_sheet = None

    # Sort the monthly sheets based on the order in the months list
    sorted_monthly_sheets = sorted(monthly_sheets, key=lambda m: months.index(m))

    # Rearrange the monthly sheets first
    for target_index, name in enumerate(sorted_monthly_sheets):
        current_index = sheet_names.index(name)
        offset = target_index - current_index
        wb.move_sheet(name, offset=offset)

    # Move the "Total" sheet to the end if it exists
    if total_sheet:
        current_index = sheet_names.index(total_sheet)
        # Move it to the last position.
        offset = len(sheet_names) - 1 - current_index  
        wb.move_sheet(total_sheet, offset=offset)

    wb.save(file_name)
    

# action = insert
def select_month(action):

    # User needs to select a month (1-12) to insert data into, else, repeat.
    while True:
        try:
            print(f"\n### {action.upper()}ER ###")
            month_number = int(input(f"\nWhich month (1-12) do you want to {action} data into? \n"))

            if not month_number in range (1, 13):
                print("\nPlease select a month from 1 to 12.")
                continue
            else:
                clear_terminal()
                # Get month name with month number.
                month_name = calendar.month_name[month_number]

                return month_number, month_name
        except ValueError:
            print("\nPlease select a month from 1 to 12.")
            continue
        except KeyboardInterrupt:
            clear_terminal()
            return



# Menu option validator
def validate_menu_option():

    MIN_OPTION = 0
    MAX_OPTION = 6
    
    # User needs to prompt 0-6, else it keeps asking to select an option.
    while True:
        try:
            menu_option = int(input(f"Select an option ({MIN_OPTION}-{MAX_OPTION}) from the menu.\n"))

            if menu_option in range (MIN_OPTION, MAX_OPTION+1):
                return menu_option
        except ValueError:
            pass
            clear_terminal()
            return None
        except UnboundLocalError:
            clear_terminal()
            return None
                


def view_data(file_name, month_name, view_yearly):


    wb = load_workbook(file_name)

    ws = wb[month_name]

    clear_terminal()
    print("\n### VIEWER ###")

    data =[]
    row_count = 0
    for row in ws.rows:
        if row_count == 0:
            row_count += 1
            continue
            
        row_values = []
    
        for cell in row:
            if isinstance(cell, openpyxl.cell.cell.MergedCell):
                continue
            row_values.append(cell.value) 

        if row_values:
            data.append(row_values)

        row_count += 1

    if month_name == "Total":
        header = ["Category", "Quantity", "Amount", "Total"]
    else:
        header = ["Day", "Description", "Amount", "Category", "Total"]

    print("\n")
    print(tabulate(data, header, tablefmt="grid"))

    if view_yearly == True:
        temp_name = file_name.replace('.xlsx', '')
        print(f'\nYearly expenses of {temp_name} were also added into a new "Total" sheet.\n')



if __name__ == "__main__":
    main()



